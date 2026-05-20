import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os

df_demand  = pd.read_csv("output/01_top_demanded_skills.csv")
df_gap     = pd.read_csv("output/02_skill_gaps.csv")
df_salary  = pd.read_csv("output/03_salary_by_skill.csv")
df_domain  = pd.read_csv("output/04_domain_analysis.csv")
df_taught  = pd.read_csv("output/05_college_skills.csv")

wb = Workbook()

DARK   = "1F2D3D"
ACCENT = "2ECC71"
HEAD   = "2C3E50"
RED    = "E74C3C"
BLUE   = "3498DB"
ORANGE = "F39C12"
WHITE  = "FFFFFF"
ALT    = "F0F4F8"
LIGHT  = "FAFBFC"

def fill(h): return PatternFill("solid", start_color=h, end_color=h)
def ft(bold=False, color="1F2D3D", size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="D5DCE4")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def header_row(ws, row, values, cols, bg=HEAD, fg=WHITE, size=10):
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col)
        c.value = val
        c.font = ft(bold=True, color=fg, size=size)
        c.fill = fill(bg)
        c.alignment = al("center")
        c.border = thin_border()
    ws.row_dimensions[row].height = 22

def data_row(ws, row, values, cols, alt=False, fmts=None):
    bg = ALT if alt else WHITE
    for i, (col, val) in enumerate(zip(cols, values)):
        c = ws.cell(row=row, column=col)
        c.value = val
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = al("center")
        if fmts and fmts[i]:
            c.number_format = fmts[i]
    ws.row_dimensions[row].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — OVERVIEW / KEY METRICS
# ══════════════════════════════════════════════════════════════════════════════
ov = wb.active
ov.title = "📊 Overview"
ov.sheet_view.showGridLines = False
set_col_widths(ov, {"A":2,"B":28,"C":18,"D":18,"E":18,"F":18,"G":2})

# Title bar
for r in range(1,4):
    for c in range(1,8):
        ov.cell(r,c).fill = fill(DARK)
ov.merge_cells("B2:F2")
t = ov["B2"]
t.value = "📈  Skill Gap Analysis — Indian Job Market vs College Curriculum"
t.font = Font(name="Arial", bold=True, color=WHITE, size=17)
t.alignment = al("center")
ov.row_dimensions[2].height = 38

# KPI cards
kpi_data = [
    ("🏢  Job Postings Analyzed",    "497",    BLUE),
    ("🎯  Unique Skills Demanded",   "20",     ACCENT),
    ("🚨  Critical Skill Gaps",      "8",      RED),
    ("📚  Skills Taught in College", "53",     ORANGE),
]
ov.merge_cells("B4:F4")
k = ov["B4"]
k.value = "PROJECT SUMMARY"
k.font = ft(bold=True, color=WHITE, size=10)
k.fill = fill(HEAD)
k.alignment = al("center")
ov.row_dimensions[4].height = 22

kpi_cols = [("B","C"),("C","D"),("D","E"),("E","F")]
for i, ((label, val, color), (c1,c2)) in enumerate(zip(kpi_data, [("B","B"),("C","C"),("D","D"),("E","E")])):
    row = 5
    col = 2 + i
    c = ov.cell(row, col)
    c.value = label
    c.font = ft(bold=True, color=WHITE, size=9)
    c.fill = fill(color)
    c.alignment = al("center", wrap=True)
    c.border = thin_border()
    ov.row_dimensions[5].height = 30

    c2 = ov.cell(6, col)
    c2.value = val
    c2.font = Font(name="Arial", bold=True, color=color, size=22)
    c2.fill = fill(WHITE)
    c2.alignment = al("center")
    c2.border = thin_border()
    ov.row_dimensions[6].height = 36

# Top demanded skills table
ov.merge_cells("B8:F8")
h2 = ov["B8"]
h2.value = "TOP SKILLS DEMANDED BY INDUSTRY"
h2.font = ft(bold=True, color=WHITE, size=10)
h2.fill = fill(HEAD)
h2.alignment = al("center")
ov.row_dimensions[8].height = 22

header_row(ov, 9, ["Skill","Category","Job Count","% of Postings","Coverage"],
           [2,3,4,5,6], bg=ACCENT, fg=WHITE)

for i, row in df_demand.head(15).iterrows():
    r = 10 + i
    cov = "✅ Taught" if row["skill"] in df_taught["skill"].values else "❌ Gap"
    data_row(ov, r,
             [row["skill"], row["category"], row["demand_count"], row["pct_jobs"]/100, cov],
             [2,3,4,5,6], alt=i%2==0,
             fmts=[None, None, "#,##0", "0.0%", None])
    ov.cell(r,2).alignment = al("left")
    ov.cell(r,2).font = ft(bold=True if cov=="❌ Gap" else False)
    color = RED if cov == "❌ Gap" else ACCENT
    ov.cell(r,6).font = ft(color=color, bold=True)

# Notes
ov.merge_cells("B26:F26")
n = ov["B26"]
n.value = "💡 Data Source: Synthetic dataset modeled on Indian tech job market (497 postings across 28 companies, 5 roles) | Analysis: SQL + Python | Dashboard: Excel"
n.font = ft(italic=True, color="888888", size=8)
n.alignment = al("center")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — SKILL GAP DETAILS
# ══════════════════════════════════════════════════════════════════════════════
sg = wb.create_sheet("🚨 Skill Gaps")
sg.sheet_view.showGridLines = False
set_col_widths(sg, {"A":2,"B":26,"C":18,"D":16,"E":20,"F":20,"G":2})

for r in range(1,4):
    for c in range(1,8):
        sg.cell(r,c).fill = fill(DARK)
sg.merge_cells("B2:F2")
t2 = sg["B2"]
t2.value = "🚨  Critical Skill Gaps — High Demand, Low College Coverage"
t2.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t2.alignment = al("center")
sg.row_dimensions[2].height = 36

header_row(sg, 4, ["Skill","Category","% Jobs Requiring","College Courses","Gap Severity"],
           [2,3,4,5,6], bg=RED, fg=WHITE)

for i, row in df_gap.iterrows():
    r = 5 + i
    severity = "🔴 Critical" if row["college_coverage"]==0 else "🟡 Moderate"
    data_row(sg, r,
             [row["skill"],row["category"],row["pct_jobs"]/100,
              int(row["college_coverage"]),severity],
             [2,3,4,5,6], alt=i%2==0,
             fmts=[None,None,"0.0%","#,##0",None])
    sg.cell(r,2).alignment = al("left")
    sg.cell(r,2).font = ft(bold=True)
    clr = RED if severity.startswith("🔴") else ORANGE
    sg.cell(r,6).font = ft(color=clr, bold=True)

# Insight box
sg.merge_cells("B21:F21")
sg["B21"].value = "INSIGHT"
sg["B21"].font = ft(bold=True, color=WHITE, size=10)
sg["B21"].fill = fill(HEAD)
sg["B21"].alignment = al("center")
sg.row_dimensions[21].height = 22

sg.merge_cells("B22:F24")
ins = sg["B22"]
ins.value = ("SQL, Power BI, Tableau, A/B Testing, and Business Intelligence top the gap list — "
             "skills that appear in 30–60% of job postings but receive zero dedicated coverage in "
             "most BCA/B.Tech curricula. These represent the most actionable self-learning priorities "
             "for students targeting Data Analyst or BI Developer roles.")
ins.font = ft(size=10, italic=True)
ins.alignment = al("left", wrap=True)
ins.fill = fill("FFF3F3")
ins.border = thin_border()
for r in range(22,25):
    sg.row_dimensions[r].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SALARY ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
sal = wb.create_sheet("💰 Salary Premium")
sal.sheet_view.showGridLines = False
set_col_widths(sal, {"A":2,"B":26,"C":18,"D":16,"E":18,"F":18,"G":2})

for r in range(1,4):
    for c in range(1,8):
        sal.cell(r,c).fill = fill(DARK)
sal.merge_cells("B2:F2")
t3 = sal["B2"]
t3.value = "💰  Salary Premium — Gap Skills vs College-Covered Skills"
t3.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t3.alignment = al("center")
sal.row_dimensions[2].height = 36

df_salary_s = df_salary.sort_values("avg_salary", ascending=False)
df_salary_s["coverage_label"] = df_salary_s["college_coverage"].apply(
    lambda x: "❌ Gap" if x==0 else ("⚠️ Partial" if x==1 else "✅ Covered")
)

header_row(sal, 4, ["Skill","Avg Salary (LPA)","Jobs Counted","Coverage","Salary Band"],
           [2,3,4,5,6], bg=BLUE, fg=WHITE)
set_col_widths(sal, {"G":18})

avg_sal = df_salary_s["avg_salary"].mean()
for i, row in df_salary_s.iterrows():
    r = 5 + list(df_salary_s.index).index(i)
    band = "⬆️ Above Avg" if row["avg_salary"] >= avg_sal else "⬇️ Below Avg"
    data_row(sal, r,
             [row["skill"],row["avg_salary"],
              int(row["job_count"]),row["coverage_label"],band],
             [2,3,4,5,6], alt=r%2==0,
             fmts=[None,"₹#,##0.0","#,##0",None,None])
    sal.cell(r,2).alignment = al("left")
    clr = RED if "Gap" in row["coverage_label"] else (ORANGE if "Partial" in row["coverage_label"] else ACCENT)
    sal.cell(r,5).font = ft(color=clr, bold=True)
    sal.cell(r,3).font = ft(bold=True, color=BLUE)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — DOMAIN ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
dom = wb.create_sheet("🏢 Domain Analysis")
dom.sheet_view.showGridLines = False
set_col_widths(dom, {"A":2,"B":22,"C":20,"D":20,"E":18,"F":16,"G":2})

for r in range(1,4):
    for c in range(1,8):
        dom.cell(r,c).fill = fill(DARK)
dom.merge_cells("B2:F2")
t4 = dom["B2"]
t4.value = "🏢  Domain-wise Skill Gap Severity"
t4.font = Font(name="Arial", bold=True, color=WHITE, size=15)
t4.alignment = al("center")
dom.row_dimensions[2].height = 36

df_domain_s = df_domain.sort_values("gap_pct", ascending=False)
header_row(dom, 4, ["Domain","Skills Demanded","Covered by College","Gap %","Avg Salary (LPA)"],
           [2,3,4,5,6], bg=ORANGE, fg=WHITE)

for i, (_, row) in enumerate(df_domain_s.iterrows()):
    r = 5 + i
    data_row(dom, r,
             [row["domain"],int(row["skills_demanded"]),
              int(row["skills_covered"]),row["gap_pct"]/100,row["avg_salary"]],
             [2,3,4,5,6], alt=i%2==0,
             fmts=[None,"#,##0","#,##0","0.0%","₹#,##0.0"])
    dom.cell(r,2).alignment = al("left")
    dom.cell(r,2).font = ft(bold=True)
    clr = RED if row["gap_pct"]>60 else (ORANGE if row["gap_pct"]>45 else ACCENT)
    dom.cell(r,5).font = ft(color=clr, bold=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — CHARTS (embed all 6 images)
# ══════════════════════════════════════════════════════════════════════════════
ch = wb.create_sheet("📉 Charts")
ch.sheet_view.showGridLines = False
for r in range(1,4):
    for c in range(1,20):
        ch.cell(r,c).fill = fill(DARK)
ch.merge_cells("A2:R2")
tc = ch["A2"]
tc.value = "📉  Visual Analysis — 6 Charts"
tc.font = Font(name="Arial", bold=True, color=WHITE, size=16)
tc.alignment = al("center")
ch.row_dimensions[2].height = 36

chart_files = sorted([f for f in os.listdir("charts") if f.endswith(".png")])
positions = ["A4","J4","A29","J29","A54","J54"]
for fname, pos in zip(chart_files, positions):
    try:
        img = XLImage(f"charts/{fname}")
        img.width  = 520
        img.height = 320
        ch.add_image(img, pos)
    except Exception as e:
        print(f"  Skipping {fname}: {e}")

wb.save("output/Skill_Gap_Analysis.xlsx")
print("Excel dashboard saved.")
